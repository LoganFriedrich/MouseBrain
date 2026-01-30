#!/usr/bin/env python3
"""
Fetch and parse eLife cervical injection reference data.

Downloads the eLife source data and extracts cervical (C4) injection counts
for comparison with our cervical injection pipeline results.

Source: Wang et al. (2022) eLife
"Brain-wide analysis of the supraspinal connectome reveals anatomical
correlates to functional recovery after spinal injury"
DOI: 10.7554/eLife.76254
"""

import urllib.request
import tempfile
from pathlib import Path
import sys

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent))


def download_elife_data():
    """Download the eLife source data Excel file."""
    url = "https://cdn.elifesciences.org/articles/76254/elife-76254-data1-v2.xlsx"

    print(f"Downloading eLife source data from:\n  {url}")

    # Download to temp file
    temp_path = Path(tempfile.gettempdir()) / "elife-76254-data1-v2.xlsx"

    try:
        urllib.request.urlretrieve(url, temp_path)
        print(f"  Downloaded to: {temp_path}")
        return temp_path
    except Exception as e:
        print(f"  ERROR: Failed to download: {e}")
        return None


def parse_cervical_data(xlsx_path: Path):
    """
    Parse the Excel file and extract cervical injection data.

    Returns dict of {region: (mean, std, n)} for cervical-projecting neurons.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return None

    print(f"\nParsing Excel file: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print(f"  Sheets found: {wb.sheetnames}")

    # Look for cervical data
    cervical_data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n  Examining sheet: {sheet_name}")

        # Print first few rows to understand structure
        print("  First 5 rows:")
        for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True)):
            # Filter None values for cleaner output
            row_clean = [str(v)[:20] if v is not None else '' for v in row[:10]]
            print(f"    Row {i+1}: {row_clean}")

        # Look for headers containing 'cervical', 'C4', or region names
        headers = None
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
            row_strs = [str(v).lower() if v else '' for v in row]
            if any('cervical' in s or 'c4' in s for s in row_strs):
                print(f"  Found cervical header at row {row_idx + 1}")
                headers = row
                break
            if any('region' in s or 'brain' in s or 'nucleus' in s for s in row_strs):
                print(f"  Found region header at row {row_idx + 1}")
                headers = row

    wb.close()
    return cervical_data


def main():
    """Main function to fetch and parse eLife cervical data."""
    print("=" * 70)
    print("eLife Cervical Reference Data Fetcher")
    print("=" * 70)

    # Download
    xlsx_path = download_elife_data()
    if not xlsx_path:
        return

    # Parse
    cervical_data = parse_cervical_data(xlsx_path)

    if cervical_data:
        print("\n" + "=" * 70)
        print("CERVICAL REFERENCE DATA (C4 injection)")
        print("=" * 70)
        for region, (mean, std, n) in sorted(cervical_data.items()):
            print(f"  {region}: {mean} +/- {std} (n={n})")
    else:
        print("\nCould not automatically parse cervical data.")
        print("The Excel file has been downloaded - please open it manually to inspect.")
        print(f"\nFile location: {xlsx_path}")
        print("\nLook for sheets/columns containing:")
        print("  - 'C4' or 'cervical' labels")
        print("  - Uninjured/control animal data")
        print("  - Region-by-region cell counts")


if __name__ == "__main__":
    main()
